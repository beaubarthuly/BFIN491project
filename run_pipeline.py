
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from utils.data_utils import (
    INITIAL_CAPITAL,
    REQUIRED_CANDIDATE_COLUMNS,
    REQUIRED_CONSTRAINT_PARAMETERS,
    REQUIRED_INHERITED_COLUMNS,
    REQUIRED_INPUT_PARAMETERS,
    WorkbookValidationError,
    build_manifest_rows,
    build_performance_shell,
    build_portfolio_summary,
    build_price_quality_summary,
    build_risk_notes,
    clean_candidate_rows,
    compute_performance_metrics,
    collect_requested_tickers,
    download_yahoo_prices,
    ensure_output_dirs,
    get_paths,
    load_cached_price_panel,
    parse_excel_date,
    plot_backtest_legacy_vs_benchmark,
    plot_inherited_drawdown,
    plot_inherited_fund_overview,
    plot_inherited_weights_snapshot,
    read_key_value_sheet,
    read_table_sheet,
    reconstruct_legacy_fund,
    save_clean_price_panels,
    save_figure,
    save_table,
    setup_logging,
    validate_candidates,
    validate_constraints,
    validate_inherited_fund,
    validate_required_columns,
    validate_required_keys,
    validate_required_sheets,
    validate_revised_portfolio,
    write_failure_status,
)
from utils.portfolio_utils import (
    build_candidate_screen,
    construct_revised_active_fund,
    construct_revised_static_fund,
    plot_candidate_recent_returns,
    plot_candidate_risk_return,
    plot_legacy_static_active_vs_benchmark,
    plot_legacy_static_vs_benchmark,
    plot_revised_active_weights,
    plot_revised_static_weights,
    prepare_revised_static_selection,
)
from utils.factor_utils import (
    build_capm_summary,
    build_ff3_summary,
    build_portfolio_return_panel,
    build_rolling_beta_table,
    load_ff3_factors,
    plot_factor_alpha_beta,
    plot_ff3_exposures,
    plot_rolling_beta,
)
from utils.risk_utils import (
    build_scenario_outputs,
    plot_monte_carlo_distribution,
    plot_stress_scenarios,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run the Fund Management Dashboard capstone pipeline."
    )
    parser.add_argument(
        "--workbook",
        type=str,
        default="fund_manager_control.xlsx",
        help="Path to the Excel control workbook.",
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Use the cached clean price panel if it exists instead of downloading fresh Yahoo data.",
    )
    return parser.parse_args()


def populate_metrics(
    performance_summary: pd.DataFrame,
    portfolio_name: str,
    metrics: dict[str, float],
) -> pd.DataFrame:
    perf = performance_summary.copy()
    mask = perf["portfolio"].eq(portfolio_name)
    for metric, value in metrics.items():
        if metric in perf.columns:
            perf.loc[mask, metric] = value
    return perf


def append_risk_note(
    risk_df: pd.DataFrame,
    item: str,
    value: str,
    notes: str,
) -> pd.DataFrame:
    extra = pd.DataFrame([{"item": item, "value": value, "notes": notes}])
    return pd.concat([risk_df, extra], ignore_index=True)


PIPELINE_VERSION = "starter_pack_v1_outputs_polish"


def _safe_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _set_metric_value(df: pd.DataFrame, portfolio_name: str, column: str, value) -> pd.DataFrame:
    out = df.copy()
    if column not in out.columns:
        return out
    mask = out["portfolio"].eq(portfolio_name)
    out.loc[mask, column] = value
    return out


def _standardize_date_df(df: pd.DataFrame, date_col: str = "date") -> pd.DataFrame:
    out = df.copy()
    if date_col in out.columns:
        out[date_col] = pd.to_datetime(out[date_col], errors="coerce")
        out = out.dropna(subset=[date_col]).sort_values(date_col)
    return out


def build_workbook_portfolio_snapshot(portfolio_summary: pd.DataFrame) -> pd.DataFrame:
    if portfolio_summary is None or portfolio_summary.empty:
        return pd.DataFrame(
            columns=[
                "ticker",
                "source",
                "decision",
                "selected_for_final",
                "static_weight",
                "latest_active_weight",
                "notes",
            ]
        )

    out = portfolio_summary.copy()
    if "keep_add_drop" in out.columns and "decision" not in out.columns:
        out = out.rename(columns={"keep_add_drop": "decision"})
    if "selected_for_final" not in out.columns:
        out["selected_for_final"] = np.where(out["ticker"].notna(), "Yes", "")
    else:
        out["selected_for_final"] = out["selected_for_final"].replace("", np.nan)
        out["selected_for_final"] = out["selected_for_final"].fillna(
            np.where(out["ticker"].notna(), "Yes", "")
        )
    for col in ["static_weight", "latest_active_weight"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    keep_cols = [
        "ticker",
        "source",
        "decision",
        "selected_for_final",
        "static_weight",
        "latest_active_weight",
        "notes",
    ]
    for col in keep_cols:
        if col not in out.columns:
            out[col] = np.nan
    return out[keep_cols].head(10).copy()


def build_workbook_performance_summary(
    legacy: dict,
    static_result: dict | None,
    active_result: dict | None,
    oos_end: pd.Timestamp,
) -> pd.DataFrame:
    base = build_performance_shell().copy()
    if "ending_value" not in base.columns:
        base.insert(1, "ending_value", np.nan)
    if "Benchmark" not in base["portfolio"].astype(str).tolist():
        base = pd.concat(
            [
                base,
                pd.DataFrame(
                    [{"portfolio": "Benchmark", "ending_value": np.nan}]
                ),
            ],
            ignore_index=True,
        )

    base = populate_metrics(base, "Legacy Fund", legacy.get("legacy_metrics", {}))
    if static_result is not None:
        base = populate_metrics(base, "Revised Static Fund", static_result.get("metrics", {}))
    if active_result is not None:
        base = populate_metrics(base, "Revised Active Fund", active_result.get("metrics", {}))

    # Legacy
    legacy_daily = _standardize_date_df(legacy["fund_daily"])
    legacy_oos_first = pd.Timestamp(legacy["oos_first_trade_date"])
    legacy_oos = legacy_daily.loc[
        (legacy_daily["date"] >= legacy_oos_first)
        & (legacy_daily["date"] <= pd.Timestamp(oos_end))
    ].copy()
    if not legacy_oos.empty:
        base = _set_metric_value(
            base,
            "Legacy Fund",
            "ending_value",
            float(pd.to_numeric(legacy_oos["legacy_fund_value"], errors="coerce").dropna().iloc[-1]),
        )

    # Static
    if static_result is not None:
        static_daily = _standardize_date_df(static_result["static_daily"])
        if not static_daily.empty:
            base = _set_metric_value(
                base,
                "Revised Static Fund",
                "ending_value",
                float(pd.to_numeric(static_daily["revised_static_value"], errors="coerce").dropna().iloc[-1]),
            )

    # Active
    if active_result is not None:
        active_daily = _standardize_date_df(active_result["active_daily"])
        if not active_daily.empty:
            base = _set_metric_value(
                base,
                "Revised Active Fund",
                "ending_value",
                float(pd.to_numeric(active_daily["revised_active_value"], errors="coerce").dropna().iloc[-1]),
            )

    # Benchmark
    benchmark_daily = _standardize_date_df(legacy["benchmark_daily"])
    if not benchmark_daily.empty:
        benchmark_returns = pd.to_numeric(benchmark_daily["benchmark_return"], errors="coerce")
        benchmark_metrics = compute_performance_metrics(benchmark_returns, turnover=0.0)
        base = populate_metrics(base, "Benchmark", benchmark_metrics)

        legacy_start_cap = np.nan
        if not legacy_oos.empty:
            legacy_start_cap = float(pd.to_numeric(legacy_oos["legacy_fund_value"], errors="coerce").iloc[0])

        benchmark_growth = pd.to_numeric(benchmark_daily["benchmark_growth_of_1"], errors="coerce").dropna()
        if benchmark_growth.shape[0] and pd.notna(legacy_start_cap):
            base = _set_metric_value(
                base,
                "Benchmark",
                "ending_value",
                float(legacy_start_cap * benchmark_growth.iloc[-1]),
            )

    preferred_order = ["Legacy Fund", "Revised Static Fund", "Revised Active Fund", "Benchmark"]
    base["portfolio"] = pd.Categorical(base["portfolio"], preferred_order, ordered=True)
    base = base.sort_values("portfolio").reset_index(drop=True)
    return base[
        [
            "portfolio",
            "ending_value",
            "total_return",
            "ann_return",
            "ann_vol",
            "sharpe",
            "sortino",
            "max_drawdown",
            "turnover",
        ]
    ].copy()


def build_workbook_active_summary(
    constraints: dict,
    active_result: dict | None,
) -> pd.DataFrame:
    last_rebalance_date = "Pending"
    n_rebalances = 0
    notes_last = "No active rebalances have been executed yet."
    if active_result is not None:
        rebalance_log = active_result.get("rebalance_log")
        if isinstance(rebalance_log, pd.DataFrame) and not rebalance_log.empty:
            last_rebalance_date = _safe_text(rebalance_log.iloc[-1].get("effective_date"))
            n_rebalances = int(active_result.get("n_rebalances", len(rebalance_log)))
            notes_last = (
                f"Latest effective rebalance date. Average monthly turnover="
                f"{float(active_result.get('avg_turnover', 0.0)):.1%}."
            )
        else:
            n_rebalances = int(active_result.get("n_rebalances", 0))

    rows = [
        {
            "field": "active_rule",
            "value": _safe_text(active_result.get("active_rule_used") if active_result else constraints.get("active_rule")),
            "notes": "Monthly rule-based rebalancing on the fixed post-2020 universe.",
        },
        {
            "field": "rebalance_frequency",
            "value": _safe_text(constraints.get("rebalance_frequency")),
            "notes": "Required project setting from the control workbook.",
        },
        {
            "field": "estimation_window_months",
            "value": pd.to_numeric(constraints.get("estimation_window_months"), errors="coerce"),
            "notes": "Rolling lookback window used by the active rule when applicable.",
        },
        {
            "field": "turnover_cap",
            "value": pd.to_numeric(constraints.get("turnover_cap"), errors="coerce"),
            "notes": "Applied when the proposed rebalance exceeds the control limit.",
        },
        {
            "field": "transaction_cost_bps",
            "value": pd.to_numeric(constraints.get("transaction_cost_bps"), errors="coerce"),
            "notes": "Linear trading-cost assumption used in the starter backtest.",
        },
        {
            "field": "last_rebalance_date",
            "value": last_rebalance_date,
            "notes": notes_last,
        },
        {
            "field": "number_of_rebalances",
            "value": n_rebalances,
            "notes": "Count of effective monthly rebalances during the out-of-sample window.",
        },
    ]
    return pd.DataFrame(rows)


def build_workbook_factor_snapshot(
    factor_capm_summary: pd.DataFrame,
    factor_rolling_beta: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "portfolio",
        "alpha_ann",
        "beta",
        "r_squared",
        "t_stat_beta",
        "p_value_beta",
        "note",
    ]
    if factor_capm_summary is None or factor_capm_summary.empty:
        return pd.DataFrame(
            [{"portfolio": "Pending", "alpha_ann": np.nan, "beta": np.nan, "r_squared": np.nan, "t_stat_beta": np.nan, "p_value_beta": np.nan, "note": "Factor/regression outputs are not ready yet."}],
            columns=columns,
        )

    roll = factor_rolling_beta.copy() if isinstance(factor_rolling_beta, pd.DataFrame) else pd.DataFrame()
    if not roll.empty and "date" in roll.columns:
        roll["date"] = pd.to_datetime(roll["date"], errors="coerce")

    rows = []
    for _, row in factor_capm_summary.iterrows():
        portfolio_name = _safe_text(row.get("portfolio"))
        latest_beta = np.nan
        if not roll.empty and {"portfolio", "rolling_beta", "date"}.issubset(roll.columns):
            tmp = roll.loc[roll["portfolio"].eq(portfolio_name)].dropna(subset=["rolling_beta", "date"])
            if not tmp.empty:
                latest_beta = pd.to_numeric(tmp.sort_values("date")["rolling_beta"], errors="coerce").dropna().iloc[-1]
        note_parts = []
        if pd.notna(latest_beta):
            note_parts.append(f"latest rolling beta={float(latest_beta):.2f}")
        market_proxy = _safe_text(row.get("market_proxy"))
        if market_proxy:
            note_parts.append(f"proxy={market_proxy}")
        n_obs = pd.to_numeric(row.get("n_obs"), errors="coerce")
        if pd.notna(n_obs):
            note_parts.append(f"n={int(n_obs)}")
        rows.append(
            {
                "portfolio": portfolio_name,
                "alpha_ann": pd.to_numeric(row.get("alpha_ann"), errors="coerce"),
                "beta": pd.to_numeric(row.get("beta"), errors="coerce"),
                "r_squared": pd.to_numeric(row.get("r_squared"), errors="coerce"),
                "t_stat_beta": pd.to_numeric(row.get("beta_tstat"), errors="coerce"),
                "p_value_beta": pd.to_numeric(row.get("beta_pvalue"), errors="coerce"),
                "note": "; ".join(note_parts) if note_parts else _safe_text(row.get("notes")),
            }
        )
    return pd.DataFrame(rows, columns=columns).head(4)


def build_workbook_scenario_snapshot(
    scenario_result: dict | None,
) -> pd.DataFrame:
    columns = [
        "portfolio",
        "mc_mean_terminal",
        "mc_p5",
        "mc_p50",
        "mc_p95",
        "worst_stress_name",
        "worst_stress_impact",
        "note",
    ]
    default_rows = []
    for portfolio_name in ["Revised Static Fund", "Revised Active Fund"]:
        default_rows.append(
            {
                "portfolio": portfolio_name,
                "mc_mean_terminal": np.nan,
                "mc_p5": np.nan,
                "mc_p50": np.nan,
                "mc_p95": np.nan,
                "worst_stress_name": "",
                "worst_stress_impact": np.nan,
                "note": "Scenario/stress outputs are not ready yet.",
            }
        )
    if scenario_result is None:
        return pd.DataFrame(default_rows, columns=columns)

    sim_summary = scenario_result.get("simulation_summary")
    sim_draws = scenario_result.get("simulation_draws")
    stress_summary = scenario_result.get("stress_summary")

    if not isinstance(sim_summary, pd.DataFrame):
        sim_summary = pd.DataFrame()
    if not isinstance(sim_draws, pd.DataFrame):
        sim_draws = pd.DataFrame()
    if not isinstance(stress_summary, pd.DataFrame):
        stress_summary = pd.DataFrame()

    rows = []
    for portfolio_name in ["Revised Static Fund", "Revised Active Fund"]:
        row = {
            "portfolio": portfolio_name,
            "mc_mean_terminal": np.nan,
            "mc_p5": np.nan,
            "mc_p50": np.nan,
            "mc_p95": np.nan,
            "worst_stress_name": "",
            "worst_stress_impact": np.nan,
            "note": "Scenario/stress outputs are not ready yet.",
        }

        if not sim_summary.empty and "portfolio" in sim_summary.columns:
            sim_row = sim_summary.loc[sim_summary["portfolio"].eq(portfolio_name)]
            if not sim_row.empty:
                sim_row = sim_row.iloc[0]
                row["mc_mean_terminal"] = pd.to_numeric(sim_row.get("mean_terminal_return"), errors="coerce")
                row["mc_p5"] = pd.to_numeric(sim_row.get("pct_05_terminal_return"), errors="coerce")
                row["mc_p50"] = pd.to_numeric(sim_row.get("median_terminal_return"), errors="coerce")
                row["note"] = f"Monte Carlo horizon={_safe_text(sim_row.get('horizon_days'))} trading days; n_sims={_safe_text(sim_row.get('n_sims'))}"

        if not sim_draws.empty and {"portfolio", "terminal_return"}.issubset(sim_draws.columns):
            draw_series = pd.to_numeric(
                sim_draws.loc[sim_draws["portfolio"].eq(portfolio_name), "terminal_return"],
                errors="coerce",
            ).dropna()
            if not draw_series.empty:
                row["mc_p95"] = float(np.nanpercentile(draw_series, 95))

        if not stress_summary.empty:
            impact_col = "static_portfolio_impact" if portfolio_name == "Revised Static Fund" else "active_portfolio_impact"
            if impact_col in stress_summary.columns:
                tmp = stress_summary.copy()
                tmp[impact_col] = pd.to_numeric(tmp[impact_col], errors="coerce")
                tmp = tmp.dropna(subset=[impact_col]).sort_values(impact_col)
                if not tmp.empty:
                    worst = tmp.iloc[0]
                    row["worst_stress_name"] = _safe_text(worst.get("scenario"))
                    row["worst_stress_impact"] = pd.to_numeric(worst.get(impact_col), errors="coerce")
                    if _safe_text(worst.get("notes")):
                        row["note"] = _safe_text(worst.get("notes"))

        rows.append(row)

    return pd.DataFrame(rows, columns=columns)


def select_major_artifacts(manifest_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    preferred = [
        "fig_inherited_fund_overview.png",
        "fig_legacy_static_active_benchmark.png",
        "fig_legacy_static_benchmark.png",
        "fig_backtest_legacy_vs_benchmark.png",
        "fig_candidate_risk_return.png",
        "fig_revised_static_weights.png",
        "fig_revised_active_weights.png",
        "fig_factor_alpha_beta.png",
        "fig_factor_rolling_beta.png",
        "fig_scenario_monte_carlo_distribution.png",
        "fig_scenario_stress_impacts.png",
        "tbl_backtest_legacy_static_active_benchmark.csv",
        "tbl_backtest_legacy_static_benchmark.csv",
        "tbl_backtest_legacy_vs_benchmark.csv",
        "tbl_factor_capm_summary.csv",
        "tbl_scenario_stress_summary.csv",
        "tbl_project_manifest.csv",
    ]
    selected = []
    seen = set()
    for filename in preferred:
        for row in manifest_rows:
            if row.get("artifact_name") == filename and filename not in seen:
                selected.append(
                    {
                        "artifact_type": row.get("artifact_type", ""),
                        "filename": filename,
                        "relative_path": row.get("relative_path", ""),
                    }
                )
                seen.add(filename)
                break
    return selected[:20]


def _clear_excel_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def _write_dataframe_rows(ws, start_row: int, start_col: int, df: pd.DataFrame, columns: list[str], max_rows: int) -> None:
    _clear_excel_range(ws, start_row, start_row + max_rows - 1, start_col, start_col + len(columns) - 1)
    if df is None or df.empty:
        return
    for row_idx, (_, row) in enumerate(df.head(max_rows).iterrows(), start=start_row):
        for col_idx, column in enumerate(columns, start=start_col):
            value = row.get(column)
            if isinstance(value, float) and np.isnan(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)


def _write_notes_manifest_sheet(ws, warnings: list[str], manifest_rows: list[dict[str, str]]) -> None:
    ws["B4"] = warnings[0] if len(warnings) >= 1 else ""
    ws["B5"] = warnings[1] if len(warnings) >= 2 else ""
    ws["B6"] = warnings[2] if len(warnings) >= 3 else ""
    start_row = 21
    _clear_excel_range(ws, min_row=start_row, max_row=max(start_row + len(manifest_rows) + 15, 120), min_col=1, max_col=6)
    for row_idx, row in enumerate(manifest_rows, start=start_row):
        ws.cell(row=row_idx, column=1, value=row.get("created_at", ""))
        ws.cell(row=row_idx, column=2, value=row.get("artifact_type", ""))
        ws.cell(row=row_idx, column=3, value=row.get("artifact_name", ""))
        ws.cell(row=row_idx, column=4, value=row.get("relative_path", ""))
        ws.cell(row=row_idx, column=5, value=row.get("status", ""))
        ws.cell(row=row_idx, column=6, value=row.get("notes", ""))


def update_outputs_workbook(
    workbook_path: Path,
    run_time: str,
    status: str,
    workbook_check: str,
    dashboard_ready: str,
    workbook_portfolio_summary: pd.DataFrame,
    workbook_performance_summary: pd.DataFrame,
    active_summary_df: pd.DataFrame,
    factor_snapshot_df: pd.DataFrame,
    scenario_snapshot_df: pd.DataFrame,
    major_artifacts: list[dict[str, str]],
    warnings: list[str],
    manifest_rows: list[dict[str, str]],
) -> None:
    wb = load_workbook(workbook_path)
    outputs_ws = wb["Outputs"]
    notes_ws = wb["Notes_Manifest"]

    outputs_ws["B3"] = run_time
    outputs_ws["B4"] = status
    outputs_ws["B5"] = workbook_check
    outputs_ws["B6"] = dashboard_ready
    outputs_ws["B7"] = PIPELINE_VERSION

    _write_dataframe_rows(
        outputs_ws,
        start_row=10,
        start_col=1,
        df=workbook_portfolio_summary,
        columns=["ticker", "source", "decision", "selected_for_final", "static_weight", "latest_active_weight", "notes"],
        max_rows=10,
    )
    _write_dataframe_rows(
        outputs_ws,
        start_row=23,
        start_col=1,
        df=workbook_performance_summary,
        columns=["portfolio", "ending_value", "total_return", "ann_return", "ann_vol", "sharpe", "sortino", "max_drawdown", "turnover"],
        max_rows=4,
    )
    _write_dataframe_rows(
        outputs_ws,
        start_row=31,
        start_col=1,
        df=active_summary_df,
        columns=["field", "value", "notes"],
        max_rows=7,
    )
    _write_dataframe_rows(
        outputs_ws,
        start_row=40,
        start_col=1,
        df=factor_snapshot_df,
        columns=["portfolio", "alpha_ann", "beta", "r_squared", "t_stat_beta", "p_value_beta", "note"],
        max_rows=4,
    )
    _write_dataframe_rows(
        outputs_ws,
        start_row=47,
        start_col=1,
        df=scenario_snapshot_df,
        columns=["portfolio", "mc_mean_terminal", "mc_p5", "mc_p50", "mc_p95", "worst_stress_name", "worst_stress_impact", "note"],
        max_rows=6,
    )
    artifact_df = pd.DataFrame(major_artifacts)
    _write_dataframe_rows(
        outputs_ws,
        start_row=56,
        start_col=1,
        df=artifact_df,
        columns=["artifact_type", "filename", "relative_path"],
        max_rows=20,
    )

    _write_notes_manifest_sheet(notes_ws, warnings, manifest_rows)
    wb.save(workbook_path)


def main() -> int:
    args = parse_args()
    paths = get_paths(__file__, args.workbook)
    logger, log_path = setup_logging(paths.logs_dir)

    if not paths.workbook_path.exists():
        logger.error("Workbook not found: %s", paths.workbook_path)
        return 1

    try:
        ensure_output_dirs(paths, logger)
        validate_required_sheets(paths.workbook_path)

        inputs = read_key_value_sheet(paths.workbook_path, "Inputs")
        constraints = read_key_value_sheet(paths.workbook_path, "Constraints")
        inherited_df = read_table_sheet(paths.workbook_path, "InheritedFund")
        candidates_df = read_table_sheet(paths.workbook_path, "Candidates")

        validate_required_keys(inputs, REQUIRED_INPUT_PARAMETERS, "Inputs")
        validate_required_keys(constraints, REQUIRED_CONSTRAINT_PARAMETERS, "Constraints")
        validate_required_columns(inherited_df, REQUIRED_INHERITED_COLUMNS, "InheritedFund")
        validate_required_columns(candidates_df, REQUIRED_CANDIDATE_COLUMNS, "Candidates")

        warnings: list[str] = []
        warnings.extend(validate_inherited_fund(inherited_df))
        warnings.extend(validate_candidates(candidates_df))
        warnings.extend(validate_constraints(constraints))
        warnings.extend(validate_revised_portfolio(inherited_df, candidates_df))

        fund_inception = parse_excel_date(inputs["fund_inception"], "Inputs.fund_inception")
        decision_date = parse_excel_date(inputs["decision_date"], "Inputs.decision_date")
        oos_start = parse_excel_date(inputs["oos_start"], "Inputs.oos_start")
        oos_end = parse_excel_date(inputs["oos_end"], "Inputs.oos_end")
        benchmark_ticker = str(inputs["market_benchmark"]).strip().upper()

        cleaned_candidates = clean_candidate_rows(candidates_df)
        requested_tickers = collect_requested_tickers(
            inherited_df=inherited_df,
            candidates_df=cleaned_candidates,
            benchmark_ticker=benchmark_ticker,
        )
        logger.info("Requested tickers: %s", ", ".join(requested_tickers))

        clean_price_path = paths.clean_dir / "prices_adjclose_daily.csv"
        if args.no_download and clean_price_path.exists():
            price_panel = load_cached_price_panel(clean_price_path, requested_tickers)
            logger.info("Loaded cached clean price panel: %s", clean_price_path)
            raw_download_path = None
        else:
            _, price_panel, raw_download_path = download_yahoo_prices(
                tickers=requested_tickers,
                start=fund_inception,
                end=oos_end,
                raw_dir=paths.raw_dir,
                logger=logger,
            )

        clean_price_path, clean_return_path, _return_panel = save_clean_price_panels(
            price_panel=price_panel,
            clean_dir=paths.clean_dir,
            logger=logger,
        )

        price_quality = build_price_quality_summary(price_panel)

        legacy = reconstruct_legacy_fund(
            price_panel=price_panel,
            inherited_df=inherited_df,
            benchmark_ticker=benchmark_ticker,
            fund_inception=fund_inception,
            decision_date=decision_date,
            oos_start=oos_start,
            oos_end=oos_end,
            initial_capital=INITIAL_CAPITAL,
        )

        logger.info("Legacy fund first trade date: %s", legacy["first_trade_date"].strftime("%Y-%m-%d"))
        logger.info("Decision snapshot date: %s", legacy["decision_snapshot_date"].strftime("%Y-%m-%d"))
        logger.info("OOS first trade date: %s", legacy["oos_first_trade_date"].strftime("%Y-%m-%d"))

        candidate_screen = build_candidate_screen(
            price_panel=price_panel,
            candidates_df=cleaned_candidates,
            decision_date=decision_date,
            benchmark_ticker=benchmark_ticker,
            legacy_fund_daily=legacy["fund_daily"],
        )

        revised_static_info = prepare_revised_static_selection(
            inherited_df=inherited_df,
            candidates_df=cleaned_candidates,
        )
        warnings.extend(revised_static_info["warnings"])

        selection_df = revised_static_info["selection"]
        if selection_df.empty:
            portfolio_summary = build_portfolio_summary(inherited_df, candidates_df)
        else:
            portfolio_summary = selection_df.copy()
            if "latest_active_weight" in portfolio_summary.columns:
                static_as_numeric = pd.to_numeric(portfolio_summary["static_weight"], errors="coerce")
                portfolio_summary["latest_active_weight"] = pd.to_numeric(
                    portfolio_summary["latest_active_weight"], errors="coerce"
                )
                if portfolio_summary["latest_active_weight"].isna().all():
                    portfolio_summary["latest_active_weight"] = static_as_numeric

        performance_summary = build_performance_shell()
        performance_summary = populate_metrics(performance_summary, "Legacy Fund", legacy["legacy_metrics"])

        risk_notes = build_risk_notes(
            inputs=inputs,
            constraints=constraints,
            portfolio_summary=portfolio_summary,
            legacy_metrics=legacy["legacy_metrics"],
            legacy_snapshot=legacy["weights_snapshot"],
        )

        if not candidate_screen.empty:
            good_candidates = candidate_screen["status"].eq("ok").sum()
            risk_notes = append_risk_note(
                risk_notes,
                item="candidate_screen_note",
                value=f"{int(good_candidates)} ready",
                notes=(
                    f"Candidate research table exported for {len(candidate_screen)} rows. "
                    "Metrics use information available through 2019-12-31 only."
                ),
            )

        static_result = None
        static_backtest_compare = pd.DataFrame()
        active_result = None
        active_backtest_compare = pd.DataFrame()
        factor_capm_summary = pd.DataFrame()
        factor_rolling_beta = pd.DataFrame()
        factor_ff3_summary = pd.DataFrame()
        scenario_result = None
        if revised_static_info["ready"]:
            static_result = construct_revised_static_fund(
                price_panel=price_panel,
                selection_df=selection_df,
                legacy_fund_daily=legacy["fund_daily"],
                benchmark_daily=legacy["benchmark_daily"],
                oos_start=oos_start,
                oos_end=oos_end,
            )
            logger.info(
                "Revised static fund first trade date: %s",
                static_result["first_trade_date"].strftime("%Y-%m-%d"),
            )
            performance_summary = populate_metrics(
                performance_summary,
                "Revised Static Fund",
                static_result["metrics"],
            )
            static_backtest_compare = static_result["compare"].copy()

            max_weight = pd.to_numeric(selection_df["static_weight"], errors="coerce").max()
            risk_notes = append_risk_note(
                risk_notes,
                item="revised_static_note",
                value=f"max weight {max_weight:.1%}",
                notes=(
                    f"Revised static fund is ready. "
                    f"Start date={static_result['first_trade_date'].strftime('%Y-%m-%d')}; "
                    f"selected names={selection_df.shape[0]}."
                ),
            )
        elif not selection_df.empty:
            risk_notes = append_risk_note(
                risk_notes,
                item="revised_static_note",
                value="Pending",
                notes=(
                    "Revised static portfolio selection has started, but the static backtest is not ready yet. "
                    "Most common causes: not exactly 10 selected names or incomplete target weights."
                ),
            )

        if static_result is not None:
            try:
                active_result = construct_revised_active_fund(
                    price_panel=price_panel,
                    selection_df=selection_df,
                    legacy_fund_daily=legacy["fund_daily"],
                    static_daily=static_result["static_daily"],
                    benchmark_daily=legacy["benchmark_daily"],
                    constraints=constraints,
                    oos_start=oos_start,
                    oos_end=oos_end,
                )
                logger.info(
                    "Revised active fund first trade date: %s",
                    active_result["first_trade_date"].strftime("%Y-%m-%d"),
                )
                performance_summary = populate_metrics(
                    performance_summary,
                    "Revised Active Fund",
                    active_result["metrics"],
                )
                active_backtest_compare = active_result["compare"].copy()

                if not portfolio_summary.empty and {"ticker", "latest_active_weight"}.issubset(portfolio_summary.columns):
                    active_snapshot = active_result["weights_snapshot"][["ticker", "latest_active_weight"]].copy()
                    portfolio_summary = portfolio_summary.drop(columns=["latest_active_weight"], errors="ignore").merge(
                        active_snapshot,
                        on="ticker",
                        how="left",
                    )

                risk_notes = append_risk_note(
                    risk_notes,
                    item="revised_active_note",
                    value=f"{active_result['active_rule_used']} / {active_result['n_rebalances']} rebals",
                    notes=(
                        f"Revised active fund is ready. Average monthly turnover={active_result['avg_turnover']:.2%}. "
                        f"The active layer uses the same fixed post-2020 universe with monthly rule-based rebalancing."
                    ),
                )
            except Exception as active_exc:
                warnings.append(f"Revised active layer could not be completed: {active_exc}")
                risk_notes = append_risk_note(
                    risk_notes,
                    item="revised_active_note",
                    value="Pending",
                    notes=(
                        "Revised active layer is still pending. "
                        f"Reason captured in pipeline warnings: {active_exc}"
                    ),
                )
        elif not portfolio_summary.empty:
            risk_notes = append_risk_note(
                risk_notes,
                item="revised_active_note",
                value="Pending",
                notes="Revised active layer will become available after the revised static portfolio is ready.",
            )


        # Factor / regression layer: use the available OOS portfolio return series versus the benchmark.
        factor_return_panel = build_portfolio_return_panel(
            legacy_fund_daily=legacy["fund_daily"],
            static_daily=None if static_result is None else static_result["static_daily"],
            active_daily=None if active_result is None else active_result["active_daily"],
            benchmark_daily=legacy["benchmark_daily"],
            oos_start=oos_start,
            oos_end=oos_end,
        )
        factor_capm_summary = build_capm_summary(
            return_panel=factor_return_panel,
            benchmark_col="benchmark_return",
            risk_free_daily=0.0,
        )
        factor_rolling_beta = build_rolling_beta_table(
            return_panel=factor_return_panel,
            benchmark_col="benchmark_return",
            window=63,
        )

        ff3_path = paths.project_root / "data" / "ff3_factors_daily.csv"
        ff3_factors = load_ff3_factors(ff3_path, oos_start=oos_start, oos_end=oos_end)
        if not ff3_factors.empty:
            factor_ff3_summary = build_ff3_summary(
                return_panel=factor_return_panel,
                ff3_factors=ff3_factors,
                benchmark_col="benchmark_return",
            )

        if not factor_capm_summary.empty:
            ready_portfolios = ", ".join(factor_capm_summary["portfolio"].astype(str).tolist())
            risk_notes = append_risk_note(
                risk_notes,
                item="factor_regression_note",
                value=f"{len(factor_capm_summary)} portfolios",
                notes=(
                    "Starter factor layer exported CAPM-style regressions using the benchmark daily return as the market proxy "
                    f"for: {ready_portfolios}. This block is meant to explain behavior, not claim prediction."
                ),
            )
        else:
            risk_notes = append_risk_note(
                risk_notes,
                item="factor_regression_note",
                value="Pending",
                notes=(
                    "Factor/regression outputs are still pending. "
                    "At least one portfolio return series and the benchmark return series must be available."
                ),
            )

        # Scenario / stress layer: Monte Carlo distribution + historical/synthetic stress scenarios.
        if static_result is not None:
            scenario_result = build_scenario_outputs(
                price_panel=price_panel,
                selection_df=selection_df,
                active_result=active_result,
                constraints=constraints,
                oos_start=oos_start,
                oos_end=oos_end,
            )
            sim_summary_df = scenario_result["simulation_summary"]
            if isinstance(sim_summary_df, pd.DataFrame) and not sim_summary_df.empty and "var_95" in sim_summary_df.columns:
                static_var = pd.to_numeric(
                    sim_summary_df.loc[sim_summary_df["portfolio"].eq("Revised Static Fund"), "var_95"],
                    errors="coerce",
                )
                active_var = pd.to_numeric(
                    sim_summary_df.loc[sim_summary_df["portfolio"].eq("Revised Active Fund"), "var_95"],
                    errors="coerce",
                )
                static_text = f"{float(static_var.iloc[0]):.1%}" if not static_var.empty and pd.notna(static_var.iloc[0]) else "N/A"
                active_text = f"{float(active_var.iloc[0]):.1%}" if not active_var.empty and pd.notna(active_var.iloc[0]) else "N/A"
                risk_notes = append_risk_note(
                    risk_notes,
                    item="scenario_stress_note",
                    value=f"VaR95 static={static_text}; active={active_text}",
                    notes=(
                        "Scenario layer exported a Monte Carlo summary and stress-scenario table using the fixed post-2020 universe. "
                        "Historical stress rows use actual selected-universe returns; synthetic rows use parallel equity shocks."
                    ),
                )
            else:
                risk_notes = append_risk_note(
                    risk_notes,
                    item="scenario_stress_note",
                    value="Pending",
                    notes="Scenario layer could not compute a simulation summary yet.",
                )
        else:
            risk_notes = append_risk_note(
                risk_notes,
                item="scenario_stress_note",
                value="Pending",
                notes="Scenario layer will become available after the revised static portfolio is ready.",
            )

        workbook_portfolio_summary = build_workbook_portfolio_snapshot(portfolio_summary)
        workbook_performance_summary = build_workbook_performance_summary(
            legacy=legacy,
            static_result=static_result,
            active_result=active_result,
            oos_end=oos_end,
        )
        active_summary_df = build_workbook_active_summary(
            constraints=constraints,
            active_result=active_result,
        )
        factor_snapshot_df = build_workbook_factor_snapshot(
            factor_capm_summary=factor_capm_summary,
            factor_rolling_beta=factor_rolling_beta,
        )
        scenario_snapshot_df = build_workbook_scenario_snapshot(
            scenario_result=scenario_result,
        )

        saved_paths: list[Path] = []

        # Dashboard core tables
        saved_paths.append(
            save_table(
                pd.DataFrame({"Parameter": list(inputs.keys()), "Value": list(inputs.values())}),
                paths.tables_dir / "tbl_inputs.csv",
                logger,
            )
        )
        saved_paths.append(
            save_table(
                pd.DataFrame({"Parameter": list(constraints.keys()), "Value": list(constraints.values())}),
                paths.tables_dir / "tbl_constraints.csv",
                logger,
            )
        )
        saved_paths.append(save_table(inherited_df, paths.tables_dir / "tbl_inherited_fund.csv", logger))
        saved_paths.append(save_table(candidates_df, paths.tables_dir / "tbl_candidates.csv", logger))
        saved_paths.append(save_table(workbook_portfolio_summary, paths.tables_dir / "tbl_portfolio_summary.csv", logger))
        saved_paths.append(save_table(workbook_performance_summary, paths.tables_dir / "tbl_performance_summary.csv", logger))
        saved_paths.append(save_table(risk_notes, paths.tables_dir / "tbl_risk_notes.csv", logger))

        # Data-layer exports
        saved_paths.append(save_table(price_quality, paths.tables_dir / "tbl_price_quality_summary.csv", logger))
        saved_paths.append(save_table(legacy["weights_snapshot"], paths.tables_dir / "tbl_legacy_weights_snapshot.csv", logger))
        saved_paths.append(save_table(legacy["fund_daily"], paths.tables_dir / "tbl_legacy_fund_daily.csv", logger))

        legacy_weights_daily = legacy["weights_daily"].reset_index().rename(columns={"index": "date"})
        saved_paths.append(save_table(legacy_weights_daily, paths.tables_dir / "tbl_legacy_weights_daily.csv", logger))

        benchmark_daily = legacy["benchmark_daily"].copy()
        saved_paths.append(save_table(benchmark_daily, paths.tables_dir / "tbl_benchmark_daily.csv", logger))

        saved_paths.append(
            save_table(legacy["backtest_compare"], paths.tables_dir / "tbl_backtest_legacy_vs_benchmark.csv", logger)
        )

        # Candidate screen exports
        saved_paths.append(
            save_table(candidate_screen, paths.tables_dir / "tbl_candidate_screen.csv", logger)
        )

        # Revised static exports (save even when pending so the dashboard can see the file)
        if static_result is not None:
            saved_paths.append(
                save_table(static_result["static_daily"], paths.tables_dir / "tbl_revised_static_daily.csv", logger)
            )
            static_weights_daily = static_result["weights_daily"].reset_index().rename(columns={"index": "date"})
            saved_paths.append(
                save_table(static_weights_daily, paths.tables_dir / "tbl_revised_static_weights_daily.csv", logger)
            )
            saved_paths.append(
                save_table(static_result["weights_snapshot"], paths.tables_dir / "tbl_revised_static_weights_snapshot.csv", logger)
            )
            saved_paths.append(
                save_table(static_backtest_compare, paths.tables_dir / "tbl_backtest_legacy_static_benchmark.csv", logger)
            )
        else:
            pending_static = pd.DataFrame(
                [{"status": revised_static_info["status"], "notes": "Revised static backtest is not ready yet."}]
            )
            saved_paths.append(
                save_table(pending_static, paths.tables_dir / "tbl_revised_static_daily.csv", logger)
            )
            saved_paths.append(
                save_table(pending_static, paths.tables_dir / "tbl_backtest_legacy_static_benchmark.csv", logger)
            )

        # Revised active exports (save even when pending so the dashboard can see the file)
        if active_result is not None:
            saved_paths.append(
                save_table(active_result["active_daily"], paths.tables_dir / "tbl_revised_active_daily.csv", logger)
            )
            saved_paths.append(
                save_table(active_result["weights_daily"], paths.tables_dir / "tbl_revised_active_weights_daily.csv", logger)
            )
            saved_paths.append(
                save_table(active_result["weights_snapshot"], paths.tables_dir / "tbl_revised_active_weights_snapshot.csv", logger)
            )
            saved_paths.append(
                save_table(active_result["rebalance_log"], paths.tables_dir / "tbl_revised_active_rebalance_log.csv", logger)
            )
            saved_paths.append(
                save_table(active_backtest_compare, paths.tables_dir / "tbl_backtest_legacy_static_active_benchmark.csv", logger)
            )
        else:
            pending_active = pd.DataFrame(
                [{"status": "pending", "notes": "Revised active backtest is not ready yet."}]
            )
            saved_paths.append(
                save_table(pending_active, paths.tables_dir / "tbl_revised_active_daily.csv", logger)
            )
            saved_paths.append(
                save_table(pending_active, paths.tables_dir / "tbl_revised_active_weights_daily.csv", logger)
            )
            saved_paths.append(
                save_table(pending_active, paths.tables_dir / "tbl_revised_active_rebalance_log.csv", logger)
            )
            saved_paths.append(
                save_table(pending_active, paths.tables_dir / "tbl_backtest_legacy_static_active_benchmark.csv", logger)
            )


        # Factor / regression exports
        if not factor_capm_summary.empty:
            saved_paths.append(
                save_table(factor_capm_summary, paths.tables_dir / "tbl_factor_capm_summary.csv", logger)
            )
        else:
            saved_paths.append(
                save_table(
                    pd.DataFrame([{"status": "pending", "notes": "Factor/regression outputs are not ready yet."}]),
                    paths.tables_dir / "tbl_factor_capm_summary.csv",
                    logger,
                )
            )

        if not factor_rolling_beta.empty:
            saved_paths.append(
                save_table(factor_rolling_beta, paths.tables_dir / "tbl_factor_rolling_beta.csv", logger)
            )
        else:
            saved_paths.append(
                save_table(
                    pd.DataFrame([{"status": "pending", "notes": "Rolling-beta outputs are not ready yet."}]),
                    paths.tables_dir / "tbl_factor_rolling_beta.csv",
                    logger,
                )
            )

        if not factor_ff3_summary.empty:
            saved_paths.append(
                save_table(factor_ff3_summary, paths.tables_dir / "tbl_factor_ff3_summary.csv", logger)
            )

        # Scenario / stress exports
        if scenario_result is not None:
            saved_paths.append(
                save_table(
                    scenario_result["simulation_summary"],
                    paths.tables_dir / "tbl_scenario_monte_carlo_summary.csv",
                    logger,
                )
            )
            saved_paths.append(
                save_table(
                    scenario_result["simulation_draws"],
                    paths.tables_dir / "tbl_scenario_monte_carlo_draws.csv",
                    logger,
                )
            )
            saved_paths.append(
                save_table(
                    scenario_result["stress_summary"],
                    paths.tables_dir / "tbl_scenario_stress_summary.csv",
                    logger,
                )
            )
        else:
            pending_scenario = pd.DataFrame([{"status": "pending", "notes": "Scenario/stress outputs are not ready yet."}])
            saved_paths.append(
                save_table(pending_scenario, paths.tables_dir / "tbl_scenario_monte_carlo_summary.csv", logger)
            )
            saved_paths.append(
                save_table(pending_scenario, paths.tables_dir / "tbl_scenario_monte_carlo_draws.csv", logger)
            )
            saved_paths.append(
                save_table(pending_scenario, paths.tables_dir / "tbl_scenario_stress_summary.csv", logger)
            )

        # Figures: inherited-fund layer
        plot_inherited_fund_overview(legacy["fund_daily"])
        saved_paths.append(save_figure(paths.figures_dir / "fig_inherited_fund_overview.png", logger))

        plot_inherited_drawdown(legacy["fund_daily"])
        saved_paths.append(save_figure(paths.figures_dir / "fig_inherited_drawdown.png", logger))

        plot_inherited_weights_snapshot(legacy["weights_snapshot"])
        saved_paths.append(save_figure(paths.figures_dir / "fig_inherited_weights_snapshot.png", logger))

        plot_backtest_legacy_vs_benchmark(legacy["backtest_compare"], benchmark_ticker)
        saved_paths.append(save_figure(paths.figures_dir / "fig_backtest_legacy_vs_benchmark.png", logger))

        # Figures: candidate and revised-static layer
        plot_candidate_risk_return(candidate_screen)
        saved_paths.append(save_figure(paths.figures_dir / "fig_candidate_risk_return.png", logger))

        plot_candidate_recent_returns(candidate_screen)
        saved_paths.append(save_figure(paths.figures_dir / "fig_candidate_recent_return.png", logger))

        plot_revised_static_weights(portfolio_summary)
        saved_paths.append(save_figure(paths.figures_dir / "fig_revised_static_weights.png", logger))

        if static_result is not None:
            plot_legacy_static_vs_benchmark(static_backtest_compare, benchmark_ticker)
            saved_paths.append(save_figure(paths.figures_dir / "fig_legacy_static_benchmark.png", logger))

        if active_result is not None:
            plot_revised_active_weights(active_result["weights_daily"])
            saved_paths.append(save_figure(paths.figures_dir / "fig_revised_active_weights.png", logger))

            plot_legacy_static_active_vs_benchmark(active_backtest_compare, benchmark_ticker)
            saved_paths.append(save_figure(paths.figures_dir / "fig_legacy_static_active_benchmark.png", logger))


        # Figures: factor / regression and scenario layers
        plot_factor_alpha_beta(factor_capm_summary)
        saved_paths.append(save_figure(paths.figures_dir / "fig_factor_alpha_beta.png", logger))

        plot_rolling_beta(factor_rolling_beta)
        saved_paths.append(save_figure(paths.figures_dir / "fig_factor_rolling_beta.png", logger))

        plot_ff3_exposures(factor_ff3_summary)
        saved_paths.append(save_figure(paths.figures_dir / "fig_factor_ff3.png", logger))

        if scenario_result is not None:
            plot_monte_carlo_distribution(scenario_result["simulation_draws"])
            saved_paths.append(save_figure(paths.figures_dir / "fig_scenario_monte_carlo_distribution.png", logger))

            plot_stress_scenarios(scenario_result["stress_summary"], legacy_daily=legacy["fund_daily"])
            saved_paths.append(save_figure(paths.figures_dir / "fig_scenario_stress_impacts.png", logger))
        else:
            # Still generate placeholder figures so the dashboard page remains stable.
            plot_monte_carlo_distribution(pd.DataFrame())
            saved_paths.append(save_figure(paths.figures_dir / "fig_scenario_monte_carlo_distribution.png", logger))

            plot_stress_scenarios(pd.DataFrame())
            saved_paths.append(save_figure(paths.figures_dir / "fig_scenario_stress_impacts.png", logger))

        # Include raw / clean data files in the manifest.
        if raw_download_path is not None:
            saved_paths.append(raw_download_path)
        saved_paths.extend([clean_price_path, clean_return_path])

        manifest_path = paths.tables_dir / "tbl_project_manifest.csv"
        manifest_rows = build_manifest_rows(paths, log_path, saved_paths + [manifest_path])
        manifest_df = pd.DataFrame(manifest_rows)
        saved_paths.append(save_table(manifest_df, manifest_path, logger))
        manifest_rows = build_manifest_rows(paths, log_path, saved_paths)

        workbook_check = (
            "OK: required sheets, parameters, columns, and price-panel quality checks were validated."
        )
        if warnings:
            workbook_check += f" Warnings: {len(warnings)}"
        if active_result is not None:
            dashboard_ready = (
                "YES: inherited-fund, candidate-screen, revised-static, revised-active, factor/regression, and scenario/stress CSVs/figures were exported successfully."
            )
        elif static_result is not None:
            dashboard_ready = (
                "PARTIAL: inherited-fund, candidate-screen, revised-static, factor/regression, and scenario/stress outputs are ready; revised-active is still pending."
            )
        else:
            dashboard_ready = (
                "PARTIAL: inherited-fund and candidate-screen outputs are ready; revised-static and revised-active are still pending."
            )

        run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        major_artifacts = select_major_artifacts(manifest_rows)

        update_outputs_workbook(
            workbook_path=paths.workbook_path,
            run_time=run_time,
            status="SUCCESS",
            workbook_check=workbook_check,
            dashboard_ready=dashboard_ready,
            workbook_portfolio_summary=workbook_portfolio_summary,
            workbook_performance_summary=workbook_performance_summary,
            active_summary_df=active_summary_df,
            factor_snapshot_df=factor_snapshot_df,
            scenario_snapshot_df=scenario_snapshot_df,
            major_artifacts=major_artifacts,
            warnings=warnings,
            manifest_rows=manifest_rows,
        )

        logger.info("Pipeline completed successfully.")
        return 0

    except Exception as exc:
        logger.exception("Pipeline failed.")
        try:
            write_failure_status(paths.workbook_path, str(exc))
        except Exception:
            logger.exception("Failed to write failure status back to workbook.")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
