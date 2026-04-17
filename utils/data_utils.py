
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


INITIAL_CAPITAL = 1_000_000.0
TRADING_DAYS_PER_YEAR = 252

REQUIRED_SHEETS = [
    "Inputs",
    "InheritedFund",
    "Candidates",
    "Constraints",
    "Outputs",
    "Notes_Manifest",
]

REQUIRED_INPUT_PARAMETERS = [
    "project_title",
    "team_name",
    "decision_date",
    "fund_inception",
    "oos_start",
    "oos_end",
    "market_benchmark",
    "price_source",
    "macro_source",
    "risk_free_proxy",
    "dashboard_default_ticker",
    "dashboard_default_page",
]

REQUIRED_CONSTRAINT_PARAMETERS = [
    "long_only",
    "fully_invested",
    "rebalance_frequency",
    "estimation_window_months",
    "max_weight",
    "min_weight",
    "turnover_cap",
    "transaction_cost_bps",
    "active_rule",
    "optimizer_objective",
    "random_seed",
]

REQUIRED_INHERITED_COLUMNS = [
    "legacy_ticker",
    "download_ticker",
    "company_name",
    "initial_weight_2010",
    "sector",
    "decision_2020",
    "keep_in_revised",
    "target_weight_2020",
    "notes",
]

REQUIRED_CANDIDATE_COLUMNS = [
    "candidate_ticker",
    "download_ticker",
    "company_name",
    "sector_theme",
    "thesis_1line",
    "screening_note",
    "add_decision",
    "selected_for_final",
    "target_weight_2020",
    "notes",
]

VALID_ADD_DECISIONS = {"", "add", "reject", "watchlist"}
VALID_SELECTED_FLAGS = {"", "yes", "no"}
VALID_KEEP_FLAGS = {"", "yes", "no", "keep", "drop"}

@dataclass
class PipelinePaths:
    project_root: Path
    workbook_path: Path
    utils_dir: Path
    data_dir: Path
    raw_dir: Path
    clean_dir: Path
    outputs_dir: Path
    figures_dir: Path
    tables_dir: Path
    logs_dir: Path


class WorkbookValidationError(Exception):
    """Raised when the workbook structure or values are invalid."""


# ---------------------------------------------------------------------------
# Project paths and logging
# ---------------------------------------------------------------------------

def get_paths(anchor_file: str | Path, workbook_arg: str) -> PipelinePaths:
    anchor_path = Path(anchor_file).resolve()
    project_root = anchor_path.parent
    utils_dir = project_root / "utils"

    workbook_path = Path(workbook_arg)
    if not workbook_path.is_absolute():
        workbook_path = project_root / workbook_path

    data_dir = project_root / "data"
    raw_dir = data_dir / "raw"
    clean_dir = data_dir / "clean"
    outputs_dir = project_root / "outputs"
    figures_dir = outputs_dir / "figures"
    tables_dir = outputs_dir / "tables"
    logs_dir = outputs_dir / "logs"

    return PipelinePaths(
        project_root=project_root,
        workbook_path=workbook_path,
        utils_dir=utils_dir,
        data_dir=data_dir,
        raw_dir=raw_dir,
        clean_dir=clean_dir,
        outputs_dir=outputs_dir,
        figures_dir=figures_dir,
        tables_dir=tables_dir,
        logs_dir=logs_dir,
    )


def setup_logging(logs_dir: Path, logger_name: str = "capstone_pipeline"):
    import logging
    import sys

    logs_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"pipeline_{timestamp}.log"

    logger = logging.getLogger(logger_name)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    logger.info("Logging initialized.")
    return logger, log_path


def ensure_output_dirs(paths: PipelinePaths, logger) -> None:
    for folder in [
        paths.data_dir,
        paths.raw_dir,
        paths.clean_dir,
        paths.outputs_dir,
        paths.figures_dir,
        paths.tables_dir,
        paths.logs_dir,
    ]:
        folder.mkdir(parents=True, exist_ok=True)
        logger.info("Verified folder: %s", folder)


# ---------------------------------------------------------------------------
# Excel / workbook helpers
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return normalize_text(value)




def _first_nonblank_text(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def parse_excel_date(value: Any, label: str) -> pd.Timestamp:
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        raise WorkbookValidationError(f"Could not parse {label} as a valid date.")
    return pd.Timestamp(ts).normalize()


def read_key_value_sheet(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, header=2, usecols="A:C")
    df.columns = [normalize_header(col) for col in df.columns]
    df = df.dropna(how="all")
    df = df[df["Parameter"].notna()].copy()

    result: dict[str, Any] = {}
    for _, row in df.iterrows():
        key = normalize_text(row["Parameter"])
        result[key] = row["Value"]
    return result


def read_table_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, header=2)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df.columns = [normalize_header(col) for col in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def validate_required_sheets(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing:
        raise WorkbookValidationError(
            f"Workbook is missing required sheets: {', '.join(missing)}"
        )


def validate_required_keys(params: dict[str, Any], required_keys: Iterable[str], label: str) -> None:
    missing = [key for key in required_keys if key not in params]
    if missing:
        raise WorkbookValidationError(
            f"{label} is missing required parameters: {', '.join(missing)}"
        )


def validate_required_columns(df: pd.DataFrame, required_columns: Iterable[str], label: str) -> None:
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise WorkbookValidationError(
            f"{label} is missing required columns: {', '.join(missing)}"
        )


def validate_inherited_fund(df: pd.DataFrame) -> list[str]:
    warnings: list[str] = []
    if len(df) != 10:
        warnings.append(
            f"InheritedFund currently has {len(df)} rows. The case background expects 10 legacy holdings."
        )

    weights = pd.to_numeric(df["initial_weight_2010"], errors="coerce")
    if weights.isna().any():
        raise WorkbookValidationError(
            "InheritedFund.initial_weight_2010 contains nonnumeric or blank values."
        )

    if not np.isclose(weights.sum(), 1.0, atol=1e-4):
        warnings.append(
            f"Initial legacy weights sum to {weights.sum():.4f}, not 1.0000."
        )

    bad_keep = sorted(
        set(
            v
            for v in df["keep_in_revised"].map(normalize_text).str.lower()
            if v not in VALID_KEEP_FLAGS
        )
    )
    if bad_keep:
        raise WorkbookValidationError(
            f"InheritedFund.keep_in_revised has invalid values: {', '.join(bad_keep)}"
        )

    return warnings


def validate_candidates(df: pd.DataFrame) -> list[str]:
    warnings: list[str] = []
    if df.empty:
        warnings.append("Candidates sheet has no candidate rows yet. Starter run will proceed.")
        return warnings

    decisions = df["add_decision"].map(normalize_text).str.lower()
    bad_decisions = sorted(set(v for v in decisions if v not in VALID_ADD_DECISIONS))
    if bad_decisions:
        raise WorkbookValidationError(
            f"Candidates.add_decision has invalid values: {', '.join(bad_decisions)}"
        )

    selected_flags = df["selected_for_final"].map(normalize_text).str.lower()
    bad_selected = sorted(set(v for v in selected_flags if v not in VALID_SELECTED_FLAGS))
    if bad_selected:
        raise WorkbookValidationError(
            f"Candidates.selected_for_final has invalid values: {', '.join(bad_selected)}"
        )

    return warnings


def validate_constraints(constraints: dict[str, Any]) -> list[str]:
    warnings: list[str] = []

    rebalance = normalize_text(constraints.get("rebalance_frequency"))
    if rebalance.lower() != "monthly":
        warnings.append(
            f"rebalancing frequency is '{rebalance}'. The required capstone design expects 'Monthly'."
        )

    for key in ["max_weight", "min_weight", "turnover_cap", "transaction_cost_bps"]:
        value = pd.to_numeric(constraints.get(key), errors="coerce")
        if pd.isna(value):
            raise WorkbookValidationError(f"Constraints.{key} must be numeric.")

    return warnings


def validate_revised_portfolio(inherited_df: pd.DataFrame, candidates_df: pd.DataFrame) -> list[str]:
    warnings: list[str] = []

    inherited_keep = inherited_df["keep_in_revised"].map(normalize_text).str.lower()
    candidate_selected = candidates_df.get("selected_for_final", pd.Series(dtype=object))
    if not candidate_selected.empty:
        candidate_selected = candidate_selected.map(normalize_text).str.lower()

    revised_started = (
        inherited_keep.isin({"yes", "keep"}).any()
        or (not candidate_selected.empty and candidate_selected.eq("yes").any())
        or pd.to_numeric(inherited_df["target_weight_2020"], errors="coerce").notna().any()
        or (
            not candidates_df.empty
            and pd.to_numeric(candidates_df["target_weight_2020"], errors="coerce").notna().any()
        )
    )

    if not revised_started:
        warnings.append(
            "No revised 2020 portfolio selections detected yet. Pipeline will default to the legacy fund summary."
        )
        return warnings

    inherited_selected = inherited_df.loc[inherited_keep.isin({"yes", "keep"})]
    candidate_selected_rows = (
        candidates_df.loc[candidate_selected.eq("yes")] if not candidate_selected.empty else pd.DataFrame()
    )
    revised_count = len(inherited_selected) + len(candidate_selected_rows)
    if revised_count != 10:
        warnings.append(
            f"Current revised portfolio selection count is {revised_count}. Final required portfolio should contain 10 stocks."
        )

    combined_weights = pd.concat(
        [
            pd.to_numeric(inherited_selected["target_weight_2020"], errors="coerce"),
            pd.to_numeric(candidate_selected_rows.get("target_weight_2020", pd.Series(dtype=float)), errors="coerce"),
        ],
        ignore_index=True,
    )
    provided_weights = combined_weights.dropna()
    if not provided_weights.empty and not np.isclose(provided_weights.sum(), 1.0, atol=1e-4):
        warnings.append(
            f"Selected revised target weights currently sum to {provided_weights.sum():.4f}, not 1.0000."
        )

    return warnings


# ---------------------------------------------------------------------------
# Candidate / portfolio summary helpers
# ---------------------------------------------------------------------------

def build_portfolio_summary(inherited_df: pd.DataFrame, candidates_df: pd.DataFrame) -> pd.DataFrame:
    inherited = inherited_df.copy()
    candidates = candidates_df.copy()

    inherited["keep_norm"] = inherited["keep_in_revised"].map(normalize_text).str.lower()
    inherited["decision_norm"] = inherited["decision_2020"].map(normalize_text).str.lower()
    inherited["target_weight_2020"] = pd.to_numeric(inherited["target_weight_2020"], errors="coerce")
    inherited["initial_weight_2010"] = pd.to_numeric(inherited["initial_weight_2010"], errors="coerce")

    if candidates.empty:
        candidates = pd.DataFrame(columns=REQUIRED_CANDIDATE_COLUMNS)

    candidates["selected_norm"] = candidates["selected_for_final"].map(normalize_text).str.lower()
    candidates["decision_norm"] = candidates["add_decision"].map(normalize_text).str.lower()
    candidates["target_weight_2020"] = pd.to_numeric(candidates["target_weight_2020"], errors="coerce")

    revised_started = (
        inherited["keep_norm"].isin({"yes", "keep"}).any()
        or candidates["selected_norm"].eq("yes").any()
        or inherited["target_weight_2020"].notna().any()
        or candidates["target_weight_2020"].notna().any()
    )

    if not revised_started:
        legacy_summary = inherited[["legacy_ticker", "initial_weight_2010", "company_name"]].copy()
        legacy_summary["source"] = "Inherited"
        legacy_summary["keep_add_drop"] = "Legacy hold"
        legacy_summary["static_weight"] = legacy_summary["initial_weight_2010"]
        legacy_summary["latest_active_weight"] = legacy_summary["static_weight"]
        legacy_summary["notes"] = "Starter default: legacy equal-weight fund."
        legacy_summary = legacy_summary.rename(columns={"legacy_ticker": "ticker"})
        return legacy_summary[
            ["ticker", "source", "keep_add_drop", "static_weight", "latest_active_weight", "notes"]
        ].head(10)

    inherited_selected = inherited.loc[inherited["keep_norm"].isin({"yes", "keep"})].copy()
    inherited_selected["ticker"] = inherited_selected["legacy_ticker"]
    inherited_selected["source"] = "Inherited"
    inherited_selected["keep_add_drop"] = "Keep"
    inherited_selected["static_weight"] = inherited_selected["target_weight_2020"]
    inherited_selected["latest_active_weight"] = inherited_selected["static_weight"]
    inherited_selected["notes"] = inherited_selected["notes"].fillna("Kept from legacy fund.")

    candidate_selected = candidates.loc[candidates["selected_norm"].eq("yes")].copy()
    if not candidate_selected.empty:
        candidate_selected["ticker"] = candidate_selected["candidate_ticker"]
        candidate_selected["source"] = "Candidate"
        candidate_selected["keep_add_drop"] = "Add"
        candidate_selected["static_weight"] = candidate_selected["target_weight_2020"]
        candidate_selected["latest_active_weight"] = candidate_selected["static_weight"]
        candidate_selected["notes"] = candidate_selected["notes"].fillna("Added to revised fund.")

    frames = [inherited_selected]
    if not candidate_selected.empty:
        frames.append(candidate_selected)

    summary = pd.concat(frames, ignore_index=True, sort=False)
    summary = summary[
        ["ticker", "source", "keep_add_drop", "static_weight", "latest_active_weight", "notes"]
    ]
    return summary.head(10)


def build_performance_shell() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "portfolio": ["Legacy Fund", "Revised Static Fund", "Revised Active Fund"],
            "total_return": [np.nan, np.nan, np.nan],
            "ann_return": [np.nan, np.nan, np.nan],
            "ann_vol": [np.nan, np.nan, np.nan],
            "sharpe": [np.nan, np.nan, np.nan],
            "sortino": [np.nan, np.nan, np.nan],
            "max_drawdown": [np.nan, np.nan, np.nan],
            "turnover": [np.nan, np.nan, np.nan],
        }
    )


def build_risk_notes(
    inputs: dict[str, Any],
    constraints: dict[str, Any],
    portfolio_summary: pd.DataFrame,
    legacy_metrics: dict[str, float] | None = None,
    legacy_snapshot: pd.DataFrame | None = None,
) -> pd.DataFrame:
    max_weight_msg = "Pending"
    if legacy_snapshot is not None and not legacy_snapshot.empty and "weight" in legacy_snapshot.columns:
        top_row = legacy_snapshot.sort_values("weight", ascending=False).iloc[0]
        max_weight_msg = f"{top_row['ticker']}: {top_row['weight']:.2%}"

    turnover_msg = "0.00"
    if legacy_metrics is not None:
        turnover_msg = f"{legacy_metrics.get('turnover', 0.0):.2f}"

    notes = [
        {
            "item": "beta_or_factor_note",
            "value": "Pending",
            "notes": "Starter pipeline only. Factor and regression analytics have not been computed yet.",
        },
        {
            "item": "scenario_note",
            "value": "Pending",
            "notes": "Starter pipeline only. Scenario analysis and stress testing are not computed yet.",
        },
        {
            "item": "concentration_note",
            "value": max_weight_msg,
            "notes": f"Legacy portfolio snapshot around the January 2020 decision date. Total names in summary: {len(portfolio_summary)}.",
        },
        {
            "item": "implementation_note",
            "value": normalize_text(constraints.get("active_rule")),
            "notes": (
                f"Benchmark={normalize_text(inputs.get('market_benchmark'))}; "
                f"legacy turnover={turnover_msg}; "
                f"rebalance={normalize_text(constraints.get('rebalance_frequency'))}; "
                f"transaction_cost_bps={constraints.get('transaction_cost_bps')}."
            ),
        },
    ]
    return pd.DataFrame(notes)


# ---------------------------------------------------------------------------
# File save helpers and manifest
# ---------------------------------------------------------------------------

def save_table(df: pd.DataFrame, path: Path, logger) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)
    logger.info("Saved table: %s", path)
    return path


def save_figure(path: Path, logger) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    logger.info("Saved figure: %s", path)
    return path


def build_manifest_rows(
    paths: PipelinePaths,
    log_path: Path,
    extra_paths: Sequence[Path] | None = None,
) -> list[dict[str, str]]:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    artifact_paths: list[tuple[str, Path]] = [
        ("log", log_path),
    ]

    if extra_paths:
        for full_path in extra_paths:
            if full_path.suffix.lower() in {".png", ".jpg", ".jpeg", ".svg", ".html"}:
                artifact_type = "figure"
            elif full_path.suffix.lower() in {".csv", ".xlsx"}:
                artifact_type = "table"
            else:
                artifact_type = "artifact"
            artifact_paths.append((artifact_type, full_path))

    rows: list[dict[str, str]] = []
    for artifact_type, full_path in artifact_paths:
        relative_path = full_path.relative_to(paths.project_root).as_posix()
        rows.append(
            {
                "created_at": timestamp,
                "artifact_type": artifact_type,
                "artifact_name": full_path.name,
                "relative_path": relative_path,
                "status": "created" if full_path.exists() else "expected",
                "notes": "",
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Workbook update helpers
# ---------------------------------------------------------------------------

def clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def write_status_block(ws, run_time: str, status: str, workbook_check: str, dashboard_ready: str) -> None:
    ws["B3"] = run_time
    ws["B4"] = status
    ws["B5"] = workbook_check
    ws["B6"] = dashboard_ready


def write_portfolio_summary(ws, summary_df: pd.DataFrame) -> None:
    clear_range(ws, min_row=10, max_row=20, min_col=1, max_col=6)
    for row_idx, row in enumerate(summary_df.itertuples(index=False), start=10):
        ws.cell(row=row_idx, column=1, value=row.ticker)
        ws.cell(row=row_idx, column=2, value=row.source)
        ws.cell(row=row_idx, column=3, value=row.keep_add_drop)
        ws.cell(row=row_idx, column=4, value=None if pd.isna(row.static_weight) else float(row.static_weight))
        ws.cell(row=row_idx, column=5, value=None if pd.isna(row.latest_active_weight) else float(row.latest_active_weight))
        ws.cell(row=row_idx, column=6, value=row.notes)


def write_performance_summary(ws, perf_df: pd.DataFrame) -> None:
    clear_range(ws, min_row=22, max_row=26, min_col=1, max_col=8)
    for row_idx, row in enumerate(perf_df.itertuples(index=False), start=22):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and np.isnan(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)


def write_risk_notes(ws, notes_df: pd.DataFrame) -> None:
    clear_range(ws, min_row=32, max_row=38, min_col=1, max_col=3)
    for row_idx, row in enumerate(notes_df.itertuples(index=False), start=32):
        ws.cell(row=row_idx, column=1, value=row.item)
        ws.cell(row=row_idx, column=2, value=row.value)
        ws.cell(row=row_idx, column=3, value=row.notes)


def write_artifacts_block(ws, manifest_rows: list[dict[str, str]]) -> None:
    start_row = 42
    clear_range(ws, min_row=start_row, max_row=max(start_row + len(manifest_rows) + 5, 70), min_col=1, max_col=3)
    for row_idx, row in enumerate(manifest_rows, start=start_row):
        ws.cell(row=row_idx, column=1, value=row["artifact_type"])
        ws.cell(row=row_idx, column=2, value=row["artifact_name"])
        ws.cell(row=row_idx, column=3, value=row["relative_path"])


def write_notes_manifest(ws, warnings: list[str], manifest_rows: list[dict[str, str]]) -> None:
    ws["B4"] = warnings[0] if len(warnings) >= 1 else ""
    ws["B5"] = warnings[1] if len(warnings) >= 2 else ""
    ws["B6"] = warnings[2] if len(warnings) >= 3 else ""

    start_row = 21
    clear_range(ws, min_row=start_row, max_row=max(start_row + len(manifest_rows) + 10, 80), min_col=1, max_col=6)
    for row_idx, row in enumerate(manifest_rows, start=start_row):
        ws.cell(row=row_idx, column=1, value=row["created_at"])
        ws.cell(row=row_idx, column=2, value=row["artifact_type"])
        ws.cell(row=row_idx, column=3, value=row["artifact_name"])
        ws.cell(row=row_idx, column=4, value=row["relative_path"])
        ws.cell(row=row_idx, column=5, value=row["status"])
        ws.cell(row=row_idx, column=6, value=row["notes"])


def update_workbook(
    workbook_path: Path,
    run_time: str,
    status: str,
    workbook_check: str,
    dashboard_ready: str,
    portfolio_summary: pd.DataFrame,
    performance_summary: pd.DataFrame,
    risk_notes: pd.DataFrame,
    warnings: list[str],
    manifest_rows: list[dict[str, str]],
) -> None:
    wb = load_workbook(workbook_path)

    outputs_ws = wb["Outputs"]
    notes_ws = wb["Notes_Manifest"]

    write_status_block(outputs_ws, run_time, status, workbook_check, dashboard_ready)
    write_portfolio_summary(outputs_ws, portfolio_summary)
    write_performance_summary(outputs_ws, performance_summary)
    write_risk_notes(outputs_ws, risk_notes)
    write_artifacts_block(outputs_ws, manifest_rows)
    write_notes_manifest(notes_ws, warnings, manifest_rows)

    wb.save(workbook_path)


def write_failure_status(workbook_path: Path, message: str) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Outputs"]
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B4"] = "FAILED"
    ws["B5"] = message
    ws["B6"] = "No dashboard data exported"
    wb.save(workbook_path)


# ---------------------------------------------------------------------------
# Data acquisition and quality checks
# ---------------------------------------------------------------------------

def _looks_like_blank_row(row: pd.Series) -> bool:
    for value in row.tolist():
        if normalize_text(value):
            return False
    return True


def _looks_like_template_placeholder(row: pd.Series) -> bool:
    ticker = _first_nonblank_text(row.get("download_ticker"), row.get("candidate_ticker")).upper()
    repeated_letter_placeholder = (
        len(ticker) == 3 and ticker.isalpha() and len(set(ticker)) == 1
    )
    if not repeated_letter_placeholder:
        return False

    meaningful_fields = [
        normalize_text(row.get("company_name")),
        normalize_text(row.get("sector_theme")),
        normalize_text(row.get("thesis_1line")),
        normalize_text(row.get("screening_note")),
        normalize_text(row.get("add_decision")),
        normalize_text(row.get("selected_for_final")),
    ]
    return not any(meaningful_fields)


def clean_candidate_rows(candidates_df: pd.DataFrame) -> pd.DataFrame:
    if candidates_df.empty:
        return candidates_df.copy()

    keep_mask = []
    for _, row in candidates_df.iterrows():
        if _looks_like_blank_row(row):
            keep_mask.append(False)
            continue
        if _looks_like_template_placeholder(row):
            keep_mask.append(False)
            continue
        keep_mask.append(True)
    return candidates_df.loc[keep_mask].reset_index(drop=True).copy()


def collect_requested_tickers(
    inherited_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
    benchmark_ticker: str,
) -> list[str]:
    tickers: list[str] = []

    for _, row in inherited_df.iterrows():
        ticker = _first_nonblank_text(row.get("download_ticker"), row.get("legacy_ticker")).upper()
        if ticker and ticker not in tickers:
            tickers.append(ticker)

    candidates_clean = clean_candidate_rows(candidates_df)
    if not candidates_clean.empty:
        for _, row in candidates_clean.iterrows():
            ticker = _first_nonblank_text(row.get("download_ticker"), row.get("candidate_ticker")).upper()
            if ticker and ticker not in tickers:
                tickers.append(ticker)

    benchmark = normalize_text(benchmark_ticker).upper()
    if benchmark and benchmark not in tickers:
        tickers.append(benchmark)

    return tickers


def _extract_adj_close_panel(raw_df: pd.DataFrame, tickers: Sequence[str]) -> pd.DataFrame:
    if raw_df.empty:
        raise ValueError("Downloaded price table is empty.")

    if isinstance(raw_df.columns, pd.MultiIndex):
        level0 = [normalize_text(v).lower().replace(" ", "") for v in raw_df.columns.get_level_values(0)]
        level1 = [normalize_text(v).upper() for v in raw_df.columns.get_level_values(1)]

        if "adjclose" in level0:
            cols = [col for col, lv0 in zip(raw_df.columns, level0) if lv0 == "adjclose"]
            px = raw_df.loc[:, cols].copy()
            px.columns = [normalize_text(col[1]).upper() for col in cols]
        elif "adjclose" in [normalize_text(v).lower().replace(" ", "") for v in raw_df.columns.get_level_values(1)]:
            cols = [
                col
                for col in raw_df.columns
                if normalize_text(col[1]).lower().replace(" ", "") == "adjclose"
            ]
            px = raw_df.loc[:, cols].copy()
            px.columns = [normalize_text(col[0]).upper() for col in cols]
        elif "close" in level0:
            cols = [col for col, lv0 in zip(raw_df.columns, level0) if lv0 == "close"]
            px = raw_df.loc[:, cols].copy()
            px.columns = [normalize_text(col[1]).upper() for col in cols]
        else:
            raise ValueError("Could not locate Adjusted Close or Close columns in the downloaded MultiIndex price table.")
    else:
        cols = {normalize_text(c).lower().replace(" ", ""): c for c in raw_df.columns}
        if "adjclose" in cols:
            px = raw_df[[cols["adjclose"]]].copy()
        elif "close" in cols:
            px = raw_df[[cols["close"]]].copy()
        else:
            raise ValueError("Could not locate Adjusted Close or Close in the downloaded price table.")
        px.columns = [normalize_text(tickers[0]).upper()]

    px.index = pd.to_datetime(px.index, errors="coerce")
    px = px[~px.index.isna()].copy()
    px = px.loc[~px.index.duplicated(keep="first")]
    px = px.sort_index()
    px.columns = [normalize_text(c).upper() for c in px.columns]

    missing_expected = [t for t in tickers if normalize_text(t).upper() not in px.columns]
    if missing_expected:
        raise ValueError(f"Downloaded price panel is missing expected tickers: {', '.join(missing_expected)}")

    return px[tickers].copy()


def download_yahoo_prices(
    tickers: Sequence[str],
    start: pd.Timestamp,
    end: pd.Timestamp,
    raw_dir: Path,
    logger,
) -> tuple[pd.DataFrame, pd.DataFrame, Path]:
    end_exclusive = (pd.Timestamp(end) + timedelta(days=1)).strftime("%Y-%m-%d")
    try:
        import yfinance as yf
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("yfinance is required for fresh downloads. Install it or rerun with --no-download after creating a cached clean price panel.") from exc

    raw_df = yf.download(
        list(tickers),
        start=pd.Timestamp(start).strftime("%Y-%m-%d"),
        end=end_exclusive,
        auto_adjust=False,
        progress=False,
        group_by="column",
        threads=True,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_path = raw_dir / f"yf_prices_raw_{timestamp}.csv"
    raw_df.to_csv(raw_path)
    logger.info("Saved raw download: %s", raw_path)

    price_panel = _extract_adj_close_panel(raw_df, list(tickers))
    return raw_df, price_panel, raw_path


def load_cached_price_panel(clean_path: Path, tickers: Sequence[str]) -> pd.DataFrame:
    if not clean_path.exists():
        raise FileNotFoundError(f"Cached clean price panel not found: {clean_path}")

    px = pd.read_csv(clean_path, index_col=0, parse_dates=True)
    px.columns = [normalize_text(c).upper() for c in px.columns]

    missing_expected = [t for t in tickers if normalize_text(t).upper() not in px.columns]
    if missing_expected:
        raise ValueError(
            f"Cached clean price panel is missing expected tickers: {', '.join(missing_expected)}"
        )

    px.index = pd.to_datetime(px.index, errors="coerce")
    px = px[~px.index.isna()].copy()
    px = px.loc[~px.index.duplicated(keep="first")]
    px = px.sort_index()
    return px[list(tickers)].copy()


def save_clean_price_panels(
    price_panel: pd.DataFrame,
    clean_dir: Path,
    logger,
) -> tuple[Path, Path, pd.DataFrame]:
    clean_dir.mkdir(parents=True, exist_ok=True)

    returns = price_panel.pct_change().dropna(how="all")

    price_path = clean_dir / "prices_adjclose_daily.csv"
    return_path = clean_dir / "returns_daily.csv"

    price_panel.to_csv(price_path)
    returns.to_csv(return_path)

    logger.info("Saved clean price panel: %s", price_path)
    logger.info("Saved clean return panel: %s", return_path)

    return price_path, return_path, returns


def build_price_quality_summary(price_panel: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    duplicate_dates = int(price_panel.index.duplicated().sum())
    index_monotonic = bool(price_panel.index.is_monotonic_increasing)

    for ticker in price_panel.columns:
        s = price_panel[ticker]
        first_valid = s.first_valid_index()
        last_valid = s.last_valid_index()
        records.append(
            {
                "ticker": ticker,
                "n_obs": int(s.shape[0]),
                "n_missing": int(s.isna().sum()),
                "pct_missing": float(s.isna().mean()),
                "first_date": price_panel.index.min().strftime("%Y-%m-%d"),
                "last_date": price_panel.index.max().strftime("%Y-%m-%d"),
                "first_valid_date": "" if first_valid is None else pd.Timestamp(first_valid).strftime("%Y-%m-%d"),
                "last_valid_date": "" if last_valid is None else pd.Timestamp(last_valid).strftime("%Y-%m-%d"),
                "index_monotonic": index_monotonic,
                "duplicate_dates": duplicate_dates,
            }
        )
    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Legacy fund reconstruction
# ---------------------------------------------------------------------------

def compute_drawdown(series: pd.Series) -> pd.Series:
    wealth = pd.Series(series).astype(float)
    running_max = wealth.cummax()
    return wealth / running_max - 1.0


def compute_annualized_return(daily_returns: pd.Series) -> float:
    daily_returns = daily_returns.dropna().astype(float)
    if daily_returns.empty:
        return np.nan
    total_growth = float((1.0 + daily_returns).prod())
    n_periods = daily_returns.shape[0]
    if total_growth <= 0 or n_periods == 0:
        return np.nan
    return total_growth ** (TRADING_DAYS_PER_YEAR / n_periods) - 1.0


def compute_annualized_vol(daily_returns: pd.Series) -> float:
    daily_returns = daily_returns.dropna().astype(float)
    if daily_returns.empty:
        return np.nan
    return float(daily_returns.std(ddof=1) * np.sqrt(TRADING_DAYS_PER_YEAR))


def compute_sortino_ratio(daily_returns: pd.Series, risk_free_annual: float = 0.0) -> float:
    daily_returns = daily_returns.dropna().astype(float)
    if daily_returns.empty:
        return np.nan

    rf_daily = risk_free_annual / TRADING_DAYS_PER_YEAR
    excess = daily_returns - rf_daily
    downside = excess[excess < 0]
    downside_std = downside.std(ddof=1) * np.sqrt(TRADING_DAYS_PER_YEAR) if len(downside) > 1 else np.nan
    ann_excess_return = compute_annualized_return(excess)
    if downside_std in (0.0, np.nan) or pd.isna(downside_std):
        return np.nan
    return ann_excess_return / downside_std


def compute_performance_metrics(daily_returns: pd.Series, turnover: float = 0.0) -> dict[str, float]:
    daily_returns = daily_returns.dropna().astype(float)
    if daily_returns.empty:
        return {
            "total_return": np.nan,
            "ann_return": np.nan,
            "ann_vol": np.nan,
            "sharpe": np.nan,
            "sortino": np.nan,
            "max_drawdown": np.nan,
            "turnover": turnover,
        }

    total_return = float((1.0 + daily_returns).prod() - 1.0)
    ann_return = compute_annualized_return(daily_returns)
    ann_vol = compute_annualized_vol(daily_returns)
    sharpe = np.nan if pd.isna(ann_vol) or ann_vol == 0 else ann_return / ann_vol
    sortino = compute_sortino_ratio(daily_returns)
    wealth = (1.0 + daily_returns).cumprod()
    max_drawdown = float(compute_drawdown(wealth).min())

    return {
        "total_return": total_return,
        "ann_return": ann_return,
        "ann_vol": ann_vol,
        "sharpe": sharpe,
        "sortino": sortino,
        "max_drawdown": max_drawdown,
        "turnover": turnover,
    }


def reconstruct_legacy_fund(
    price_panel: pd.DataFrame,
    inherited_df: pd.DataFrame,
    benchmark_ticker: str,
    fund_inception: pd.Timestamp,
    decision_date: pd.Timestamp,
    oos_start: pd.Timestamp,
    oos_end: pd.Timestamp,
    initial_capital: float = INITIAL_CAPITAL,
) -> dict[str, Any]:
    holdings = inherited_df.copy()
    holdings["ticker"] = holdings.apply(lambda r: _first_nonblank_text(r.get("download_ticker"), r.get("legacy_ticker")).upper(), axis=1)
    holdings["initial_weight_2010"] = pd.to_numeric(holdings["initial_weight_2010"], errors="coerce")

    missing_holdings = [t for t in holdings["ticker"] if t not in price_panel.columns]
    if missing_holdings:
        raise ValueError(
            f"Price panel is missing inherited-fund tickers: {', '.join(missing_holdings)}"
        )

    price_panel = price_panel.copy().sort_index()
    holdings_px = price_panel[holdings["ticker"].tolist()].copy()

    # Find first common date after inception with all inherited prices available.
    eligible = holdings_px.loc[holdings_px.index >= fund_inception]
    start_mask = eligible.notna().all(axis=1)
    if not start_mask.any():
        raise ValueError("Could not find a common first trade date with all inherited-fund prices available.")
    first_trade_date = eligible.index[start_mask][0]

    start_prices = holdings_px.loc[first_trade_date]
    shares = (initial_capital * holdings["initial_weight_2010"].values) / start_prices.values
    shares_series = pd.Series(shares, index=holdings["ticker"].tolist(), name="shares")

    position_values = holdings_px.mul(shares_series, axis=1)
    total_value = position_values.sum(axis=1)
    daily_returns = total_value.pct_change().fillna(0.0)
    weights_daily = position_values.div(total_value, axis=0)

    fund_daily = pd.DataFrame(
        {
            "date": total_value.index,
            "legacy_fund_value": total_value.values,
            "legacy_fund_return": daily_returns.values,
            "legacy_growth_of_1": ((1.0 + daily_returns).cumprod()).values,
            "legacy_drawdown": compute_drawdown((1.0 + daily_returns).cumprod()).values,
        }
    )

    # Decision snapshot: last trading day on or before decision date.
    snapshot_candidates = total_value.loc[total_value.index <= decision_date]
    if snapshot_candidates.empty:
        raise ValueError("No legacy-fund dates available on or before the decision date.")
    snapshot_date = snapshot_candidates.index[-1]
    weights_snapshot = weights_daily.loc[snapshot_date].reset_index()
    weights_snapshot.columns = ["ticker", "weight"]
    weights_snapshot["snapshot_date"] = snapshot_date.strftime("%Y-%m-%d")

    # Out-of-sample benchmark comparison.
    oos_candidates = total_value.loc[(total_value.index >= oos_start) & (total_value.index <= oos_end)]
    if oos_candidates.empty:
        raise ValueError("No legacy-fund observations fall inside the out-of-sample window.")
    oos_first_trade = oos_candidates.index[0]

    benchmark = normalize_text(benchmark_ticker).upper()
    if benchmark not in price_panel.columns:
        raise ValueError(f"Benchmark ticker {benchmark} is missing from the price panel.")

    benchmark_px = price_panel[benchmark].dropna().loc[price_panel.index <= oos_end]
    benchmark_px = benchmark_px.loc[benchmark_px.index >= oos_first_trade]
    if benchmark_px.empty:
        raise ValueError("No benchmark observations fall inside the out-of-sample window.")

    legacy_oos = fund_daily.set_index("date").loc[oos_first_trade:oos_end].copy()
    legacy_oos["legacy_oos_return"] = legacy_oos["legacy_fund_value"].pct_change().fillna(0.0)
    legacy_oos["legacy_growth_of_1_oos"] = (1.0 + legacy_oos["legacy_oos_return"]).cumprod()
    legacy_oos["legacy_drawdown_oos"] = compute_drawdown(legacy_oos["legacy_growth_of_1_oos"])

    benchmark_daily = pd.DataFrame(index=benchmark_px.index)
    benchmark_daily["benchmark_price"] = benchmark_px
    benchmark_daily["benchmark_return"] = benchmark_daily["benchmark_price"].pct_change().fillna(0.0)
    benchmark_daily["benchmark_growth_of_1"] = (1.0 + benchmark_daily["benchmark_return"]).cumprod()
    benchmark_daily["benchmark_drawdown"] = compute_drawdown(benchmark_daily["benchmark_growth_of_1"])

    backtest_compare = legacy_oos[["legacy_growth_of_1_oos", "legacy_drawdown_oos"]].join(
        benchmark_daily[["benchmark_growth_of_1", "benchmark_drawdown"]],
        how="inner",
    )
    backtest_compare = backtest_compare.reset_index().rename(columns={"index": "date"})

    legacy_metrics = compute_performance_metrics(legacy_oos["legacy_oos_return"], turnover=0.0)

    return {
        "first_trade_date": pd.Timestamp(first_trade_date),
        "decision_snapshot_date": pd.Timestamp(snapshot_date),
        "oos_first_trade_date": pd.Timestamp(oos_first_trade),
        "shares": shares_series,
        "position_values": position_values,
        "weights_daily": weights_daily,
        "weights_snapshot": weights_snapshot,
        "fund_daily": fund_daily,
        "benchmark_daily": benchmark_daily.reset_index().rename(columns={"index": "date"}),
        "backtest_compare": backtest_compare,
        "legacy_metrics": legacy_metrics,
    }


# ---------------------------------------------------------------------------
# Plots
# ---------------------------------------------------------------------------

def plot_inherited_fund_overview(fund_daily: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(fund_daily["date"], fund_daily["legacy_fund_value"], linewidth=1.8)
    ax.set_title("Inherited Fund Value Path")
    ax.set_xlabel("Date")
    ax.set_ylabel("Portfolio value ($)")
    ax.grid(alpha=0.25)
    fig.autofmt_xdate()


def plot_inherited_drawdown(fund_daily: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(fund_daily["date"], fund_daily["legacy_drawdown"], linewidth=1.5)
    ax.set_title("Inherited Fund Drawdown")
    ax.set_xlabel("Date")
    ax.set_ylabel("Drawdown")
    ax.axhline(0.0, color="black", linewidth=0.8)
    ax.grid(alpha=0.25)
    fig.autofmt_xdate()


def plot_inherited_weights_snapshot(weights_snapshot: pd.DataFrame) -> None:
    snapshot = weights_snapshot.sort_values("weight", ascending=False)
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(snapshot["ticker"], snapshot["weight"])
    date_label = snapshot["snapshot_date"].iloc[0] if not snapshot.empty else "snapshot"
    ax.set_title(f"Inherited Fund Weights on {date_label}")
    ax.set_xlabel("Ticker")
    ax.set_ylabel("Weight")
    ax.tick_params(axis="x", rotation=45)
    ax.grid(axis="y", alpha=0.25)


def plot_backtest_legacy_vs_benchmark(backtest_compare: pd.DataFrame, benchmark_ticker: str) -> None:
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(backtest_compare["date"], backtest_compare["legacy_growth_of_1_oos"], label="Legacy Fund", linewidth=1.7)
    ax.plot(
        backtest_compare["date"],
        backtest_compare["benchmark_growth_of_1"],
        label=normalize_text(benchmark_ticker).upper(),
        linewidth=1.5,
    )
    ax.set_title("Legacy Fund vs Benchmark (Out-of-Sample)")
    ax.set_xlabel("Date")
    ax.set_ylabel("Growth of $1")
    ax.legend()
    ax.grid(alpha=0.25)
    fig.autofmt_xdate()
